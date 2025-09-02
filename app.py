from telegram_alerts import send_telegram_alert
from weapon_detection import WeaponDetector


import csv
import copy
import argparse
import itertools
from collections import Counter
from collections import deque

import cv2 as cv
import numpy as np
import mediapipe as mp # type: ignore

import cv2 as c
face_classifier = c.CascadeClassifier(c.data.haarcascades + "haarcascade_frontalface_default.xml")
face_snapshot_counter = 10
import os


from utils import CvFpsCalc
from model import KeyPointClassifier
from model import PointHistoryClassifier
import openpyxl # type: ignore
from openpyxl import Workbook # type: ignore
wb = openpyxl.load_workbook('data.xlsx')
# workbook = Workbook()
# workbook.save(filename="data.xlsx")
cap = cv.VideoCapture(4)
wb = openpyxl.Workbook()
sheet =wb.active
x1 = sheet.cell(row=1, column=1)
x1.value = 0
x2 = sheet.cell(row=1, column=2)
x2.value = 0
wb.save("data.xlsx")




def get_args():
    parser = argparse.ArgumentParser()

    parser.add_argument("--device", type=int, default=0)
    parser.add_argument("--width", help='cap width', type=int, default=960)
    parser.add_argument("--height", help='cap height', type=int, default=540)
    # parser.add_argument("--width", help='cap width', type=int, default=640)
    # parser.add_argument("--height", help='cap height', type=int, default=360)

    parser.add_argument('--use_static_image_mode', action='store_true')
    parser.add_argument("--min_detection_confidence",
                        help='min_detection_confidence',
                        type=float,
                        default=0.7)
    parser.add_argument("--min_tracking_confidence",
                        help='min_tracking_confidence',
                        type=int,
                        default=0.5)

    args = parser.parse_args()

    return args


def main():
    weapon_detector = WeaponDetector()

    # Argument parsing #################################################################
    args = get_args()

    cap_device = args.device
    cap_width = args.width
    cap_height = args.height

    use_static_image_mode = args.use_static_image_mode
    min_detection_confidence = args.min_detection_confidence
    min_tracking_confidence = args.min_tracking_confidence

    use_brect = True

    # Camera preparation ###############################################################
    cap = cv.VideoCapture(cap_device)
    cap.set(cv.CAP_PROP_FRAME_WIDTH, cap_width)
    cap.set(cv.CAP_PROP_FRAME_HEIGHT, cap_height)

    # Model load #############################################################
    mp_hands = mp.solutions.hands
    hands = mp_hands.Hands(
        static_image_mode=use_static_image_mode,
        max_num_hands=2,
        min_detection_confidence=min_detection_confidence,
        min_tracking_confidence=min_tracking_confidence,
    )

    keypoint_classifier = KeyPointClassifier()

    point_history_classifier = PointHistoryClassifier()

    # Read labels ###########################################################
    with open('model/keypoint_classifier/keypoint_classifier_label.csv',
              encoding='utf-8-sig') as f:
        keypoint_classifier_labels = csv.reader(f)
        keypoint_classifier_labels = [
            row[0] for row in keypoint_classifier_labels
        ]
    with open(
            'model/point_history_classifier/point_history_classifier_label.csv',
            encoding='utf-8-sig') as f:
        point_history_classifier_labels = csv.reader(f)
        point_history_classifier_labels = [
            row[0] for row in point_history_classifier_labels
        ]

    # FPS Measurement ########################################################
    cvFpsCalc = CvFpsCalc(buffer_len=10)

    # Coordinate history #################################################################
    history_length = 16
    point_history = deque(maxlen=history_length)

    # Finger gesture history ################################################
    finger_gesture_history = deque(maxlen=history_length)

    #  ########################################################################
    mode = 0

    while True:
        fps = cvFpsCalc.get()

        # Process Key (ESC: end) #################################################
        key = cv.waitKey(10)
        if key == 27:  # ESC
            break
        number, mode = select_mode(key, mode)

        # Camera capture #####################################################
        ret, image = cap.read()
        image = weapon_detector.detect_and_save(image)

        if not ret:
            break
        image = cv.flip(image, 1)  # Mirror display
        debug_image = copy.deepcopy(image)

        # Detection implementation #############################################################
        image = cv.cvtColor(image, cv.COLOR_BGR2RGB)

        image.flags.writeable = False
        results = hands.process(image)
        image.flags.writeable = True

        #  ####################################################################
        if results.multi_hand_landmarks is not None:
            for hand_landmarks, handedness in zip(results.multi_hand_landmarks,
                                                  results.multi_handedness):

                # Bounding box calculation
                brect = calc_bounding_rect(debug_image, hand_landmarks)
                # Landmark calculation
                landmark_list = calc_landmark_list(debug_image, hand_landmarks)
                print(landmark_list)
                # Conversion to relative coordinates / normalized coordinates
                pre_processed_landmark_list = pre_process_landmark(
                    landmark_list)
                pre_processed_point_history_list = pre_process_point_history(
                    debug_image, point_history)
                # Write to the dataset file
                logging_csv(number, mode, pre_processed_landmark_list,
                            pre_processed_point_history_list)

                # Hand sign classification
                hand_sign_id = keypoint_classifier(pre_processed_landmark_list)

                if hand_sign_id == 2:  # Point gesture

                        x1 = sheet.cell(row=1, column=1)
                        x1.value = 1

                        wb.save("data.xlsx")





                if hand_sign_id == 3:  # Point gesture
                        
                        x1 = sheet.cell(row=1, column=2)
                        x1.value = 1

                        wb.save("data.xlsx")

                        if not os.path.exists('data'):
                            os.makedirs('data')
                        if not os.path.exists('face'):
                            os.makedirs('face')
                        # count=0
                        # last_face_filename = None
                        #     suc, frame = cap.read()
                        #     if count%20==0:
                        #         cv.imwrite(f'./data/frame{count}.jpg', frame)
                        #         def send_telegram_alert(message, image_path=None):
                        #             try:
                        #                 send_telegram_alert("Distress Call Detected", image_path=filename)
                        #             except Exception as e:
                        #                 print(f"Telegram alert failed: {e}")
                        #         detect_and_save_faces(frame)
                        #     count += 1



                        count = 0
                        last_face_filename = None

                        while count != 140:
                            suc, frame = cap.read()
                            if count % 20 == 0:
                                frame_path = f'./data/frame{count}.jpg'
                                cv.imwrite(frame_path, frame)
                                last_face_filename = frame_path
                                 # Save face image
                                detect_and_save_faces(frame)  # This updates face_snapshot_counter

        # Save the last face image name for alert
                                 

                            count += 1

# Send Telegram alert once after all frames are processed
                        try:
                            if last_face_filename and os.path.exists(last_face_filename):
                                send_telegram_alert("🚨 Distress Call Detected", image_path=last_face_filename)
                            else:
                                send_telegram_alert("🚨 Distress Call Detected (no face snapshot)")
                        except Exception as e:
                            print(f"Telegram alert failed: {e}")









                # Finger gesture classification
                finger_gesture_id = 0
                point_history_len = len(pre_processed_point_history_list)
                if point_history_len == (history_length * 2):
                    finger_gesture_id = point_history_classifier(
                        pre_processed_point_history_list)

                # Calculates the gesture IDs in the latest detection
                finger_gesture_history.append(finger_gesture_id)
                most_common_fg_id = Counter(
                    finger_gesture_history).most_common()

                #Drawing part
                debug_image = draw_bounding_rect(use_brect, debug_image, brect)
                debug_image = draw_landmarks(debug_image, landmark_list)
                debug_image = draw_info_text(
                    debug_image,
                    brect,
                    handedness,
                    keypoint_classifier_labels[hand_sign_id],
                    point_history_classifier_labels[most_common_fg_id[0][0]],
                )
        else:
            point_history.append([0, 0])

        debug_image = draw_point_history(debug_image, point_history)
        debug_image = draw_info(debug_image, fps, mode, number)

        # Screen reflection #############################################################
        cv.imshow('Hand Gesture Recognition', debug_image)

    cap.release()
    cv.destroyAllWindows()


def select_mode(key, mode):
    number = -1
    if 48 <= key <= 57:  # 0 ~ 9
        number = key - 48
    if key == 110:  # n
        mode = 0
    if key == 107:  # k
        mode = 1
    if key == 104:  # h
        mode = 2
    return number, mode


def calc_bounding_rect(image, landmarks):
    image_width, image_height = image.shape[1], image.shape[0]

    landmark_array = np.empty((0, 2), int)

    for _, landmark in enumerate(landmarks.landmark):
        landmark_x = min(int(landmark.x * image_width), image_width - 1)
        landmark_y = min(int(landmark.y * image_height), image_height - 1)

        landmark_point = [np.array((landmark_x, landmark_y))]

        landmark_array = np.append(landmark_array, landmark_point, axis=0)

    x, y, w, h = cv.boundingRect(landmark_array)

    return [x, y, x + w, y + h]


def calc_landmark_list(image, landmarks):
    image_width, image_height = image.shape[1], image.shape[0]

    landmark_point = []

    # Keypoint
    for _, landmark in enumerate(landmarks.landmark):
        landmark_x = min(int(landmark.x * image_width), image_width - 1)
        landmark_y = min(int(landmark.y * image_height), image_height - 1)
        # landmark_z = landmark.z

        landmark_point.append([landmark_x, landmark_y])

    return landmark_point


def pre_process_landmark(landmark_list):
    temp_landmark_list = copy.deepcopy(landmark_list)

    # Convert to relative coordinates
    base_x, base_y = 0, 0
    for index, landmark_point in enumerate(temp_landmark_list):
        if index == 0:
            base_x, base_y = landmark_point[0], landmark_point[1]

        temp_landmark_list[index][0] = temp_landmark_list[index][0] - base_x
        temp_landmark_list[index][1] = temp_landmark_list[index][1] - base_y

    # Convert to a one-dimensional list
    temp_landmark_list = list(
        itertools.chain.from_iterable(temp_landmark_list))

    # Normalization
    max_value = max(list(map(abs, temp_landmark_list)))


    def normalize_(n):
        return n / max_value

    temp_landmark_list = list(map(normalize_, temp_landmark_list))

    return temp_landmark_list


def pre_process_point_history(image, point_history):
    image_width, image_height = image.shape[1], image.shape[0]

    temp_point_history = copy.deepcopy(point_history)

    # Convert to relative coordinates
    base_x, base_y = 0, 0
    for index, point in enumerate(temp_point_history):
        if index == 0:
            base_x, base_y = point[0], point[1]

        temp_point_history[index][0] = (temp_point_history[index][0] -
                                        base_x) / image_width
        temp_point_history[index][1] = (temp_point_history[index][1] -
                                        base_y) / image_height

    # Convert to a one-dimensional list
    temp_point_history = list(
        itertools.chain.from_iterable(temp_point_history))

    return temp_point_history


def logging_csv(number, mode, landmark_list, point_history_list):
    if mode == 0:
        pass
    if mode == 1 and (0 <= number <= 9):
        csv_path = 'model/keypoint_classifier/keypoint.csv'
        with open(csv_path, 'a', newline="") as f:
            writer = csv.writer(f)
            writer.writerow([number, *landmark_list])
    if mode == 2 and (0 <= number <= 9):
        csv_path = 'model/point_history_classifier/point_history.csv'
        with open(csv_path, 'a', newline="") as f:
            writer = csv.writer(f)
            writer.writerow([number, *point_history_list])
    return


def draw_landmarks(image, landmark_point):
    if len(landmark_point) > 0:
        # Thumb
        cv.line(image, tuple(landmark_point[2]), tuple(landmark_point[3]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[2]), tuple(landmark_point[3]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[3]), tuple(landmark_point[4]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[3]), tuple(landmark_point[4]),
                (255, 255, 255), 2)

        # Index finger
        cv.line(image, tuple(landmark_point[5]), tuple(landmark_point[6]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[5]), tuple(landmark_point[6]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[6]), tuple(landmark_point[7]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[6]), tuple(landmark_point[7]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[7]), tuple(landmark_point[8]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[7]), tuple(landmark_point[8]),
                (255, 255, 255), 2)

        # Middle finger
        cv.line(image, tuple(landmark_point[9]), tuple(landmark_point[10]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[9]), tuple(landmark_point[10]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[10]), tuple(landmark_point[11]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[10]), tuple(landmark_point[11]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[11]), tuple(landmark_point[12]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[11]), tuple(landmark_point[12]),
                (255, 255, 255), 2)

        # Ring finger
        cv.line(image, tuple(landmark_point[13]), tuple(landmark_point[14]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[13]), tuple(landmark_point[14]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[14]), tuple(landmark_point[15]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[14]), tuple(landmark_point[15]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[15]), tuple(landmark_point[16]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[15]), tuple(landmark_point[16]),
                (255, 255, 255), 2)

        # Little finger
        cv.line(image, tuple(landmark_point[17]), tuple(landmark_point[18]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[17]), tuple(landmark_point[18]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[18]), tuple(landmark_point[19]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[18]), tuple(landmark_point[19]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[19]), tuple(landmark_point[20]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[19]), tuple(landmark_point[20]),
                (255, 255, 255), 2)

        # Palm
        cv.line(image, tuple(landmark_point[0]), tuple(landmark_point[1]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[0]), tuple(landmark_point[1]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[1]), tuple(landmark_point[2]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[1]), tuple(landmark_point[2]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[2]), tuple(landmark_point[5]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[2]), tuple(landmark_point[5]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[5]), tuple(landmark_point[9]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[5]), tuple(landmark_point[9]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[9]), tuple(landmark_point[13]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[9]), tuple(landmark_point[13]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[13]), tuple(landmark_point[17]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[13]), tuple(landmark_point[17]),
                (255, 255, 255), 2)
        cv.line(image, tuple(landmark_point[17]), tuple(landmark_point[0]),
                (0, 0, 0), 6)
        cv.line(image, tuple(landmark_point[17]), tuple(landmark_point[0]),
                (255, 255, 255), 2)

    # Key Points
    for index, landmark in enumerate(landmark_point):
        if index == 0:  # 手首1
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 1:  # 手首2
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 2:  # 親指：付け根
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 3:  # 親指：第1関節
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 4:  # 親指：指先
            cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)
        if index == 5:  # 人差指：付け根
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 6:  # 人差指：第2関節
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 7:  # 人差指：第1関節
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 8:  # 人差指：指先
            cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)
        if index == 9:  # 中指：付け根
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 10:  # 中指：第2関節
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 11:  # 中指：第1関節
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 12:  # 中指：指先
            cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)
        if index == 13:  # 薬指：付け根
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 14:  # 薬指：第2関節
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 15:  # 薬指：第1関節
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 16:  # 薬指：指先
            cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)
        if index == 17:  # 小指：付け根
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 18:  # 小指：第2関節
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 19:  # 小指：第1関節
            cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
        if index == 20:  # 小指：指先
            cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255),
                      -1)
            cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)

    return image


def draw_bounding_rect(use_brect, image, brect):
    if use_brect:
        # Outer rectangle
        cv.rectangle(image, (brect[0], brect[1]), (brect[2], brect[3]),
                     (0, 0, 0), 1)

    return image


def draw_info_text(image, brect, handedness, hand_sign_text,
                   finger_gesture_text):
    cv.rectangle(image, (brect[0], brect[1]), (brect[2], brect[1] - 22),
                 (0, 0, 0), -1)

    info_text = handedness.classification[0].label[0:]
    if hand_sign_text != "":
        info_text = info_text + ':' + hand_sign_text
    cv.putText(image, info_text, (brect[0] + 5, brect[1] - 4),
               cv.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1, cv.LINE_AA)

    if finger_gesture_text != "":
        cv.putText(image, "Finger Gesture:" + finger_gesture_text, (10, 60),
                   cv.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 0), 4, cv.LINE_AA)
        cv.putText(image, "Finger Gesture:" + finger_gesture_text, (10, 60),
                   cv.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2,
                   cv.LINE_AA)

    return image


def draw_point_history(image, point_history):
    for index, point in enumerate(point_history):
        if point[0] != 0 and point[1] != 0:
            cv.circle(image, (point[0], point[1]), 1 + int(index / 2),
                      (152, 251, 152), 2)

    return image


def draw_info(image, fps, mode, number):
    cv.putText(image, "FPS:" + str(fps), (10, 30), cv.FONT_HERSHEY_SIMPLEX,
               1.0, (0, 0, 0), 4, cv.LINE_AA)
    cv.putText(image, "FPS:" + str(fps), (10, 30), cv.FONT_HERSHEY_SIMPLEX,
               1.0, (255, 255, 255), 2, cv.LINE_AA)

    mode_string = ['Logging Key Point', 'Logging Point History']
    if 1 <= mode <= 2:
        cv.putText(image, "MODE:" + mode_string[mode - 1], (10, 90),
                   cv.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1,
                   cv.LINE_AA)
        if 0 <= number <= 9:
            cv.putText(image, "NUM:" + str(number), (10, 110),
                       cv.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1,
                       cv.LINE_AA)
    return image

def detect_and_save_faces(image, count_start=10):
    global face_snapshot_counter
    gray = c.cvtColor(image, c.COLOR_BGR2GRAY)
    faces = face_classifier.detectMultiScale(gray, 1.1, 5, minSize=(40, 40))

    if not os.path.exists("data/faces"):
        os.makedirs("data/faces")

    for (x, y, w, h) in faces:
        face = image[y:y + h, x:x + w]
        filename = f"data/faces/a{face_snapshot_counter}.jpg"
        c.imwrite(filename, face)
        face_snapshot_counter += 10



if __name__ == '__main__':
    main()


# from telegram_alerts import send_telegram_alert
# from weapon_detection import WeaponDetector
# import phone_config  # Import our phone camera configuration file

# import csv
# import copy
# import argparse
# import itertools
# from collections import Counter
# from collections import deque

# import cv2 as cv
# import numpy as np
# import mediapipe as mp # type: ignore

# import cv2 as c
# face_classifier = c.CascadeClassifier(c.data.haarcascades + "haarcascade_frontalface_default.xml")
# face_snapshot_counter = 10
# import os

# from utils import CvFpsCalc
# from model import KeyPointClassifier
# from model import PointHistoryClassifier
# import openpyxl # type: ignore
# from openpyxl import Workbook # type: ignore

# # Initialize Excel workbook
# wb = openpyxl.Workbook()
# sheet = wb.active
# x1 = sheet.cell(row=1, column=1)
# x1.value = 0
# x2 = sheet.cell(row=1, column=2)
# x2.value = 0
# wb.save("data.xlsx")

# def get_args():
#     parser = argparse.ArgumentParser()

#     parser.add_argument("--device", type=int, default=0, 
#                        help='Fallback camera device (if stream fails)')
#     parser.add_argument("--width", help='cap width', type=int, default=960)
#     parser.add_argument("--height", help='cap height', type=int, default=540)

#     parser.add_argument('--use_static_image_mode', action='store_true')
#     parser.add_argument("--min_detection_confidence",
#                         help='min_detection_confidence',
#                         type=float,
#                         default=0.7)
#     parser.add_argument("--min_tracking_confidence",
#                         help='min_tracking_confidence',
#                         type=int,
#                         default=0.5)

#     args = parser.parse_args()
#     return args

# def main():
#     weapon_detector = WeaponDetector()

#     # Argument parsing
#     args = get_args()

#     # Use phone stream URL directly from config
#     stream_url = phone_config.STREAM_URL
#     cap_device = args.device
#     cap_width = args.width
#     cap_height = args.height

#     use_static_image_mode = args.use_static_image_mode
#     min_detection_confidence = args.min_detection_confidence
#     min_tracking_confidence = args.min_tracking_confidence

#     use_brect = True

#     # Camera preparation - Automatically use HTTP stream from phone
#     print("🔍 Starting detection server...")
#     print(f"📱 Connecting to phone camera: {stream_url}")
    
#     cap = cv.VideoCapture(stream_url)
    
#     # Set buffer size to reduce latency for better real-time detection
#     cap.set(cv.CAP_PROP_BUFFERSIZE, 1)
    
#     # Check if HTTP stream connection is successful
#     if not cap.isOpened():
#         print(f"❌ Failed to connect to HTTP stream: {stream_url}")
#         print("📋 Troubleshooting:")
#         print("   1. Make sure your phone and computer are on the same Wi-Fi network")
#         print("   2. Verify the IP camera app is running on your phone")
#         print("   3. Check if the IP address and port are correct")
#         print("   4. Try accessing the URL in your browser first")
#         print(f"\n🔄 Attempting fallback to local camera...")
        
#         cap = cv.VideoCapture(cap_device)
#         if not cap.isOpened():
#             print("❌ Error: Could not open any camera source!")
#             print("💡 Please check your camera connections and try again.")
#             return
#         else:
#             print("✅ Connected to local camera as fallback")
#     else:
#         print(f"✅ Successfully connected to phone camera!")
#         print("🎯 Detection server is now running...")
#         print("📹 Weapon detection and gesture recognition active")
#         print("⌨️  Press ESC to quit")

#     # Set camera properties for optimal detection
#     cap.set(cv.CAP_PROP_FRAME_WIDTH, cap_width)
#     cap.set(cv.CAP_PROP_FRAME_HEIGHT, cap_height)

#     # Model load
#     mp_hands = mp.solutions.hands
#     hands = mp_hands.Hands(
#         static_image_mode=use_static_image_mode,
#         max_num_hands=2,
#         min_detection_confidence=min_detection_confidence,
#         min_tracking_confidence=min_tracking_confidence,
#     )

#     keypoint_classifier = KeyPointClassifier()
#     point_history_classifier = PointHistoryClassifier()

#     # Read labels
#     with open('model/keypoint_classifier/keypoint_classifier_label.csv',
#               encoding='utf-8-sig') as f:
#         keypoint_classifier_labels = csv.reader(f)
#         keypoint_classifier_labels = [
#             row[0] for row in keypoint_classifier_labels
#         ]
#     with open(
#             'model/point_history_classifier/point_history_classifier_label.csv',
#             encoding='utf-8-sig') as f:
#         point_history_classifier_labels = csv.reader(f)
#         point_history_classifier_labels = [
#             row[0] for row in point_history_classifier_labels
#         ]

#     # FPS Measurement
#     cvFpsCalc = CvFpsCalc(buffer_len=10)

#     # Coordinate history
#     history_length = 16
#     point_history = deque(maxlen=history_length)

#     # Finger gesture history
#     finger_gesture_history = deque(maxlen=history_length)

#     mode = 0

#     while True:
#         fps = cvFpsCalc.get()

#         # Process Key (ESC: end)
#         key = cv.waitKey(10)
#         if key == 27:  # ESC
#             break
#         number, mode = select_mode(key, mode)

#         # Camera capture
#         ret, image = cap.read()
        
#         if not ret:
#             print("Failed to grab frame from camera")
#             break
            
#         # Apply weapon detection first
#         image = weapon_detector.detect_and_save(image)
        
#         image = cv.flip(image, 1)  # Mirror display
#         debug_image = copy.deepcopy(image)

#         # Detection implementation
#         image = cv.cvtColor(image, cv.COLOR_BGR2RGB)
#         image.flags.writeable = False
#         results = hands.process(image)
#         image.flags.writeable = True

#         if results.multi_hand_landmarks is not None:
#             for hand_landmarks, handedness in zip(results.multi_hand_landmarks,
#                                                   results.multi_handedness):

#                 # Bounding box calculation
#                 brect = calc_bounding_rect(debug_image, hand_landmarks)
#                 # Landmark calculation
#                 landmark_list = calc_landmark_list(debug_image, hand_landmarks)
#                 print(landmark_list)
#                 # Conversion to relative coordinates / normalized coordinates
#                 pre_processed_landmark_list = pre_process_landmark(
#                     landmark_list)
#                 pre_processed_point_history_list = pre_process_point_history(
#                     debug_image, point_history)
#                 # Write to the dataset file
#                 logging_csv(number, mode, pre_processed_landmark_list,
#                             pre_processed_point_history_list)

#                 # Hand sign classification
#                 hand_sign_id = keypoint_classifier(pre_processed_landmark_list)

#                 if hand_sign_id == 2:  # Point gesture
#                     x1 = sheet.cell(row=1, column=1)
#                     x1.value = 1
#                     wb.save("data.xlsx")

#                 if hand_sign_id == 3:  # Distress gesture
#                     x1 = sheet.cell(row=1, column=2)
#                     x1.value = 1
#                     wb.save("data.xlsx")

#                     if not os.path.exists('data'):
#                         os.makedirs('data')
#                     if not os.path.exists('face'):
#                         os.makedirs('face')

#                     count = 0
#                     last_face_filename = None

#                     while count != 140:
#                         suc, frame = cap.read()
#                         if not suc:
#                             break
                            
#                         if count % 20 == 0:
#                             frame_path = f'./data/frame{count}.jpg'
#                             cv.imwrite(frame_path, frame)
#                             last_face_filename = frame_path
#                             # Save face image
#                             detect_and_save_faces(frame)

#                         count += 1

#                     # Send Telegram alert once after all frames are processed
#                     try:
#                         if last_face_filename and os.path.exists(last_face_filename):
#                             send_telegram_alert("🚨 Distress Call Detected", image_path=last_face_filename)
#                         else:
#                             send_telegram_alert("🚨 Distress Call Detected (no face snapshot)")
#                     except Exception as e:
#                         print(f"Telegram alert failed: {e}")

#                 # Finger gesture classification
#                 finger_gesture_id = 0
#                 point_history_len = len(pre_processed_point_history_list)
#                 if point_history_len == (history_length * 2):
#                     finger_gesture_id = point_history_classifier(
#                         pre_processed_point_history_list)

#                 # Calculates the gesture IDs in the latest detection
#                 finger_gesture_history.append(finger_gesture_id)
#                 most_common_fg_id = Counter(
#                     finger_gesture_history).most_common()

#                 # Drawing part
#                 debug_image = draw_bounding_rect(use_brect, debug_image, brect)
#                 debug_image = draw_landmarks(debug_image, landmark_list)
#                 debug_image = draw_info_text(
#                     debug_image,
#                     brect,
#                     handedness,
#                     keypoint_classifier_labels[hand_sign_id],
#                     point_history_classifier_labels[most_common_fg_id[0][0]],
#                 )
#         else:
#             point_history.append([0, 0])

#         debug_image = draw_point_history(debug_image, point_history)
#         debug_image = draw_info(debug_image, fps, mode, number)

#         # Screen reflection
#         cv.imshow('Hand Gesture Recognition', debug_image)

#     cap.release()
#     cv.destroyAllWindows()

# def select_mode(key, mode):
#     number = -1
#     if 48 <= key <= 57:  # 0 ~ 9
#         number = key - 48
#     if key == 110:  # n
#         mode = 0
#     if key == 107:  # k
#         mode = 1
#     if key == 104:  # h
#         mode = 2
#     return number, mode

# def calc_bounding_rect(image, landmarks):
#     image_width, image_height = image.shape[1], image.shape[0]

#     landmark_array = np.empty((0, 2), int)

#     for _, landmark in enumerate(landmarks.landmark):
#         landmark_x = min(int(landmark.x * image_width), image_width - 1)
#         landmark_y = min(int(landmark.y * image_height), image_height - 1)

#         landmark_point = [np.array((landmark_x, landmark_y))]

#         landmark_array = np.append(landmark_array, landmark_point, axis=0)

#     x, y, w, h = cv.boundingRect(landmark_array)

#     return [x, y, x + w, y + h]

# def calc_landmark_list(image, landmarks):
#     image_width, image_height = image.shape[1], image.shape[0]

#     landmark_point = []

#     # Keypoint
#     for _, landmark in enumerate(landmarks.landmark):
#         landmark_x = min(int(landmark.x * image_width), image_width - 1)
#         landmark_y = min(int(landmark.y * image_height), image_height - 1)

#         landmark_point.append([landmark_x, landmark_y])

#     return landmark_point

# def pre_process_landmark(landmark_list):
#     temp_landmark_list = copy.deepcopy(landmark_list)

#     # Convert to relative coordinates
#     base_x, base_y = 0, 0
#     for index, landmark_point in enumerate(temp_landmark_list):
#         if index == 0:
#             base_x, base_y = landmark_point[0], landmark_point[1]

#         temp_landmark_list[index][0] = temp_landmark_list[index][0] - base_x
#         temp_landmark_list[index][1] = temp_landmark_list[index][1] - base_y

#     # Convert to a one-dimensional list
#     temp_landmark_list = list(
#         itertools.chain.from_iterable(temp_landmark_list))

#     # Normalization
#     max_value = max(list(map(abs, temp_landmark_list)))

#     def normalize_(n):
#         return n / max_value

#     temp_landmark_list = list(map(normalize_, temp_landmark_list))

#     return temp_landmark_list

# def pre_process_point_history(image, point_history):
#     image_width, image_height = image.shape[1], image.shape[0]

#     temp_point_history = copy.deepcopy(point_history)

#     # Convert to relative coordinates
#     base_x, base_y = 0, 0
#     for index, point in enumerate(temp_point_history):
#         if index == 0:
#             base_x, base_y = point[0], point[1]

#         temp_point_history[index][0] = (temp_point_history[index][0] -
#                                         base_x) / image_width
#         temp_point_history[index][1] = (temp_point_history[index][1] -
#                                         base_y) / image_height

#     # Convert to a one-dimensional list
#     temp_point_history = list(
#         itertools.chain.from_iterable(temp_point_history))

#     return temp_point_history

# def logging_csv(number, mode, landmark_list, point_history_list):
#     if mode == 0:
#         pass
#     if mode == 1 and (0 <= number <= 9):
#         csv_path = 'model/keypoint_classifier/keypoint.csv'
#         with open(csv_path, 'a', newline="") as f:
#             writer = csv.writer(f)
#             writer.writerow([number, *landmark_list])
#     if mode == 2 and (0 <= number <= 9):
#         csv_path = 'model/point_history_classifier/point_history.csv'
#         with open(csv_path, 'a', newline="") as f:
#             writer = csv.writer(f)
#             writer.writerow([number, *point_history_list])
#     return

# def draw_landmarks(image, landmark_point):
#     if len(landmark_point) > 0:
#         # Thumb
#         cv.line(image, tuple(landmark_point[2]), tuple(landmark_point[3]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[2]), tuple(landmark_point[3]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[3]), tuple(landmark_point[4]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[3]), tuple(landmark_point[4]),
#                 (255, 255, 255), 2)

#         # Index finger
#         cv.line(image, tuple(landmark_point[5]), tuple(landmark_point[6]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[5]), tuple(landmark_point[6]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[6]), tuple(landmark_point[7]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[6]), tuple(landmark_point[7]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[7]), tuple(landmark_point[8]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[7]), tuple(landmark_point[8]),
#                 (255, 255, 255), 2)

#         # Middle finger
#         cv.line(image, tuple(landmark_point[9]), tuple(landmark_point[10]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[9]), tuple(landmark_point[10]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[10]), tuple(landmark_point[11]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[10]), tuple(landmark_point[11]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[11]), tuple(landmark_point[12]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[11]), tuple(landmark_point[12]),
#                 (255, 255, 255), 2)

#         # Ring finger
#         cv.line(image, tuple(landmark_point[13]), tuple(landmark_point[14]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[13]), tuple(landmark_point[14]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[14]), tuple(landmark_point[15]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[14]), tuple(landmark_point[15]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[15]), tuple(landmark_point[16]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[15]), tuple(landmark_point[16]),
#                 (255, 255, 255), 2)

#         # Little finger
#         cv.line(image, tuple(landmark_point[17]), tuple(landmark_point[18]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[17]), tuple(landmark_point[18]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[18]), tuple(landmark_point[19]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[18]), tuple(landmark_point[19]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[19]), tuple(landmark_point[20]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[19]), tuple(landmark_point[20]),
#                 (255, 255, 255), 2)

#         # Palm
#         cv.line(image, tuple(landmark_point[0]), tuple(landmark_point[1]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[0]), tuple(landmark_point[1]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[1]), tuple(landmark_point[2]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[1]), tuple(landmark_point[2]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[2]), tuple(landmark_point[5]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[2]), tuple(landmark_point[5]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[5]), tuple(landmark_point[9]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[5]), tuple(landmark_point[9]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[9]), tuple(landmark_point[13]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[9]), tuple(landmark_point[13]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[13]), tuple(landmark_point[17]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[13]), tuple(landmark_point[17]),
#                 (255, 255, 255), 2)
#         cv.line(image, tuple(landmark_point[17]), tuple(landmark_point[0]),
#                 (0, 0, 0), 6)
#         cv.line(image, tuple(landmark_point[17]), tuple(landmark_point[0]),
#                 (255, 255, 255), 2)

#     # Key Points
#     for index, landmark in enumerate(landmark_point):
#         if index == 0:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 1:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 2:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 3:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 4:
#             cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)
#         if index == 5:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 6:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 7:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 8:
#             cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)
#         if index == 9:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 10:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 11:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 12:
#             cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)
#         if index == 13:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 14:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 15:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 16:
#             cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)
#         if index == 17:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 18:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 19:
#             cv.circle(image, (landmark[0], landmark[1]), 5, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 5, (0, 0, 0), 1)
#         if index == 20:
#             cv.circle(image, (landmark[0], landmark[1]), 8, (255, 255, 255), -1)
#             cv.circle(image, (landmark[0], landmark[1]), 8, (0, 0, 0), 1)

#     return image

# def draw_bounding_rect(use_brect, image, brect):
#     if use_brect:
#         # Outer rectangle
#         cv.rectangle(image, (brect[0], brect[1]), (brect[2], brect[3]),
#                      (0, 0, 0), 1)

#     return image

# def draw_info_text(image, brect, handedness, hand_sign_text,
#                    finger_gesture_text):
#     cv.rectangle(image, (brect[0], brect[1]), (brect[2], brect[1] - 22),
#                  (0, 0, 0), -1)

#     info_text = handedness.classification[0].label[0:]
#     if hand_sign_text != "":
#         info_text = info_text + ':' + hand_sign_text
#     cv.putText(image, info_text, (brect[0] + 5, brect[1] - 4),
#                cv.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1, cv.LINE_AA)

#     if finger_gesture_text != "":
#         cv.putText(image, "Finger Gesture:" + finger_gesture_text, (10, 60),
#                    cv.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 0), 4, cv.LINE_AA)
#         cv.putText(image, "Finger Gesture:" + finger_gesture_text, (10, 60),
#                    cv.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2,
#                    cv.LINE_AA)

#     return image

# def draw_point_history(image, point_history):
#     for index, point in enumerate(point_history):
#         if point[0] != 0 and point[1] != 0:
#             cv.circle(image, (point[0], point[1]), 1 + int(index / 2),
#                       (152, 251, 152), 2)

#     return image

# def draw_info(image, fps, mode, number):
#     cv.putText(image, "FPS:" + str(fps), (10, 30), cv.FONT_HERSHEY_SIMPLEX,
#                1.0, (0, 0, 0), 4, cv.LINE_AA)
#     cv.putText(image, "FPS:" + str(fps), (10, 30), cv.FONT_HERSHEY_SIMPLEX,
#                1.0, (255, 255, 255), 2, cv.LINE_AA)

#     mode_string = ['Logging Key Point', 'Logging Point History']
#     if 1 <= mode <= 2:
#         cv.putText(image, "MODE:" + mode_string[mode - 1], (10, 90),
#                    cv.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1,
#                    cv.LINE_AA)
#         if 0 <= number <= 9:
#             cv.putText(image, "NUM:" + str(number), (10, 110),
#                        cv.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1,
#                        cv.LINE_AA)
#     return image

# def detect_and_save_faces(image, count_start=10):
#     global face_snapshot_counter
#     gray = c.cvtColor(image, c.COLOR_BGR2GRAY)
#     faces = face_classifier.detectMultiScale(gray, 1.1, 5, minSize=(40, 40))

#     if not os.path.exists("data/faces"):
#         os.makedirs("data/faces")

#     for (x, y, w, h) in faces:
#         face = image[y:y + h, x:x + w]
#         filename = f"data/faces/a{face_snapshot_counter}.jpg"
#         c.imwrite(filename, face)
#         face_snapshot_counter += 10

# if __name__ == '__main__':
#     main()